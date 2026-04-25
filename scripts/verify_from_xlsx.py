"""
Compare DB transactions for a date against Ken's xlsx.

Usage:
  python scripts/verify_from_xlsx.py <YYYY-MM-DD> <xlsx_path> [prev_xlsx_path]

  YYYY-MM-DD    : date to verify
  xlsx_path     : this day's xlsx
  prev_xlsx_path: previous day's xlsx (needed to strip carry-in rows accurately)

Output: per-currency BUY/SELL count + qty comparison, any THAN=0 flags.

Examples:
  python scripts/verify_from_xlsx.py 2026-04-21 excel_files/4.21.26.xlsx excel_files/4.20.26.xlsx
  python scripts/verify_from_xlsx.py 2026-04-22 excel_files/4.22.26.xlsx
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from datetime import date
from collections import defaultdict
import openpyxl

from app.core.database import SessionLocal
from app.models.transaction import Transaction, TxnType
from app.models.currency import DailyPosition

CODE_MAP   = {'NTD': 'TWD', 'QR': 'QAR', 'KD': 'KWD'}
SKIP_CODES = {'BHD', 'BND', 'OMR'}

PAIRS_2ND = [
    ('AED',1,2), ('AUD',3,4), ('CAD',5,6), ('CHF',7,8),
    ('CNY',9,10), ('EUR',11,12), ('GBP',13,14), ('HKD',15,16),
    ('MYR',17,18), ('NTD',19,20), ('NZD',21,22), ('QR',23,24),
    ('SAR',25,26), ('SGD',27,28), ('THB',29,30),
]
PAIRS_OTHERS = [
    ('BHD',1,2), ('BND',3,4), ('DKK',5,6), ('IDR',7,8),
    ('INR',9,10), ('MOP',11,12), ('NOK',13,14), ('OMR',15,16),
    ('TYR',17,18), ('VND',19,20), ('KD',21,22),
]


def norm(code):
    return CODE_MAP.get(str(code).strip(), str(code).strip())


def is_num(v):
    return isinstance(v, (int, float)) and v > 0


def ws(wb, name):
    """Return worksheet tolerating case/spacing/typo variants."""
    if name in wb.sheetnames:
        return wb[name]
    def _n(s):
        return ' '.join(s.lower().split())
    n = _n(name)
    for s in wb.sheetnames:
        if _n(s) == n:
            return wb[s]
    n2 = n.replace('buy', 'bu')
    for s in wb.sheetnames:
        if _n(s).replace('buy', 'bu') == n2:
            return wb[s]
    raise KeyError(f"Sheet '{name}' not found. Available: {wb.sheetnames}")


def parse_main(sheet, stop_at_gap=True):
    pairs = [('USD', 2, 3), ('JPY', 5, 6), ('KRW', 8, 9)]
    txns, empty = [], 0
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        found = False
        for code, qc, rc in pairs:
            q = row[qc] if len(row) > qc else None
            r = row[rc] if len(row) > rc else None
            if is_num(q) and is_num(r):
                txns.append((code, q, r))
                found = True
            # No EXCESS for MAIN currencies (USD/JPY/KRW always require a rate)
        if stop_at_gap:
            empty = 0 if found else empty + 1
            if empty >= 2:
                break
    return txns


def parse_pairs(sheet, pairs, data_start, stop_at_gap=True):
    txns, empty = [], 0
    for row in list(sheet.iter_rows(values_only=True))[data_start:]:
        found = False
        for code, qc, rc in pairs:
            if code in SKIP_CODES:
                continue
            q = row[qc] if len(row) > qc else None
            r = row[rc] if len(row) > rc else None
            if is_num(q) and is_num(r):
                txns.append((norm(code), q, r))
                found = True
            elif is_num(q) and not is_num(r):
                txns.append((norm(code), q, None))
                found = True
        if stop_at_gap:
            empty = 0 if found else empty + 1
            if empty >= 2:
                break
    return txns


def read_stocksleft(wb):
    stocks = {}
    for row in wb['STOCKSLEFT'].iter_rows(values_only=True):
        code, qty, rate = row[0], row[1], row[2]
        if not isinstance(code, str) or not (2 <= len(code.strip()) <= 5):
            continue
        db_code = CODE_MAP.get(code.strip(), code.strip())
        if db_code in SKIP_CODES:
            continue
        if isinstance(qty, (int, float)) and qty > 0 and isinstance(rate, (int, float)) and rate > 0:
            stocks[db_code] = (qty, rate)
    return stocks


def main():
    args = sys.argv[1:]
    if len(args) < 2:
        print("Usage: verify_from_xlsx.py <YYYY-MM-DD> <xlsx_path> [prev_xlsx_path]")
        sys.exit(1)

    target_date = date.fromisoformat(args[0])
    xlsx_path   = args[1]
    prev_xlsx   = args[2] if len(args) > 2 else None

    print(f"\n{'='*60}")
    print(f"  VERIFY {target_date}  vs  {xlsx_path}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Use stop=False only when prev_xlsx is given (avoids early exit on BUY sheets)
    stop = prev_xlsx is None

    raw_buys = (
        parse_main(ws(wb, 'BUY x MAIN'), stop)
      + parse_pairs(ws(wb, 'BUY x 2ND'),    PAIRS_2ND,    2, stop)
      + parse_pairs(ws(wb, 'BUY  x OTHERS'), PAIRS_OTHERS, 2, stop)
    )
    raw_sells = (
        parse_main(ws(wb, 'SELL x MAIN'), True)
      + parse_pairs(ws(wb, 'SELL  x 2ND'),   PAIRS_2ND,    1, True)
      + parse_pairs(ws(wb, 'SELL x OTHERS'), PAIRS_OTHERS, 1, True)
    )

    # Determine carry-in set for stripping
    if prev_xlsx:
        prev_wb = openpyxl.load_workbook(prev_xlsx, data_only=True)
        carry_in = read_stocksleft(prev_wb)
        prev_wb.close()
        carry_in_set = {(code, qty, rate) for code, (qty, rate) in carry_in.items()}
    else:
        db = SessionLocal()
        try:
            pos_rows = db.query(DailyPosition).filter_by(date=target_date).all()
            carry_in_set = {(p.currency_code, p.carry_in_qty, p.carry_in_rate) for p in pos_rows}
        finally:
            db.close()

    # Strip carry-in rows from buys
    buys_clean = [(c, q, r) for c, q, r in raw_buys if (c, q, r) not in carry_in_set]
    stripped = len(raw_buys) - len(buys_clean)
    if stripped:
        print(f"\n  Stripped {stripped} carry-in row(s) from Excel BUYs")

    wb.close()

    # Aggregate Excel data by currency
    xlsx_buys:  dict[str, list] = defaultdict(list)
    xlsx_sells: dict[str, list] = defaultdict(list)
    for code, qty, rate in buys_clean:
        xlsx_buys[code].append((qty, rate))
    for code, qty, rate in raw_sells:
        xlsx_sells[code].append((qty, rate))

    # Fetch DB data
    db = SessionLocal()
    try:
        db_txns = db.query(Transaction).filter(Transaction.date == target_date).all()
    finally:
        db.close()

    db_buys:  dict[str, list] = defaultdict(list)
    db_sells: dict[str, list] = defaultdict(list)
    for t in db_txns:
        if t.type in (TxnType.BUY, TxnType.EXCESS):
            db_buys[t.currency_code].append((t.foreign_amt, t.rate, t.type))
        elif t.type == TxnType.SELL:
            db_sells[t.currency_code].append((t.foreign_amt, t.rate, t.than))

    all_codes = sorted(set(xlsx_buys) | set(xlsx_sells) | set(db_buys) | set(db_sells))

    # ── BUY comparison ────────────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print("  BUYS (excluding carry-in)")
    print(f"{'─'*60}")
    print(f"  {'Currency':<8} {'Excel #':>8} {'Excel qty':>12} {'DB #':>8} {'DB qty':>12}  Status")
    buy_ok = buy_issues = 0
    for code in all_codes:
        ex_rows = xlsx_buys.get(code, [])
        db_rows = db_buys.get(code, [])
        ex_cnt  = len(ex_rows)
        db_cnt  = len(db_rows)
        ex_qty  = round(sum(q for q, _ in ex_rows), 2)
        db_qty  = round(sum(q for q, _, _ in db_rows), 2)
        if not ex_rows and not db_rows:
            continue
        ok = (ex_cnt == db_cnt and abs(ex_qty - db_qty) < 0.01)
        flag = "OK" if ok else "MISMATCH <<"
        if ok:
            buy_ok += 1
        else:
            buy_issues += 1
        print(f"  {code:<8} {ex_cnt:>8} {ex_qty:>12.2f} {db_cnt:>8} {db_qty:>12.2f}  {flag}")

    print(f"\n  Buys: {buy_ok} OK, {buy_issues} mismatch(es)")

    # ── SELL comparison ───────────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print("  SELLS")
    print(f"{'─'*60}")
    print(f"  {'Currency':<8} {'Excel #':>8} {'Excel qty':>12} {'DB #':>8} {'DB qty':>12}  {'THAN=0?':<10} Status")
    sell_ok = sell_issues = 0
    for code in all_codes:
        ex_rows = xlsx_sells.get(code, [])
        db_rows = db_sells.get(code, [])
        ex_cnt  = len(ex_rows)
        db_cnt  = len(db_rows)
        ex_qty  = round(sum(q for q, _ in ex_rows), 2)
        db_qty  = round(sum(q for q, _, _ in db_rows), 2)
        if not ex_rows and not db_rows:
            continue
        zero_than = [q for q, _, th in db_rows if th == 0]
        ok = (ex_cnt == db_cnt and abs(ex_qty - db_qty) < 0.01)
        flag = "OK" if ok else "MISMATCH <<"
        than_flag = f"THAN=0 ({len(zero_than)})" if zero_than else ""
        if ok:
            sell_ok += 1
        else:
            sell_issues += 1
        print(f"  {code:<8} {ex_cnt:>8} {ex_qty:>12.2f} {db_cnt:>8} {db_qty:>12.2f}  {than_flag:<14} {flag}")

    print(f"\n  Sells: {sell_ok} OK, {sell_issues} mismatch(es)")

    # ── Summary ───────────────────────────────────────────────────────────────
    total_than = round(sum(t.than for t in db_txns if t.type == TxnType.SELL), 2)
    print(f"\n{'─'*60}")
    print(f"  Total DB THAN for {target_date}: ₱{total_than:,.2f}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
