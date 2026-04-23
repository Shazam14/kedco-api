"""
Seed April 5, 2026 data from /root/projects/4.5.26.xlsx

  - DailyRate   : derived from transaction rates (modal buy / actual sell)
  - DailyPosition : carry-in from April 4 STOCKSLEFT
  - Transaction : all BUY/SELL sheets, fabricated timestamps

Cashier: jas (AM 08:00–13:00), ana (PM 13:00–17:30)
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from datetime import date
from collections import Counter
import openpyxl

from app.core.database import SessionLocal
from app.models.currency import DailyRate, DailyPosition
from app.models.transaction import Transaction, TxnType, TxnSource

TARGET_DATE = date(2026, 4, 5)
DATE_SUFFIX = "0405"
XLSX        = "/root/projects/4.5.26.xlsx"

CODE_MAP   = {'NTD': 'TWD', 'QR': 'QAR', 'KD': 'KWD'}
SKIP_CODES = {'TYR', 'MOP', 'BHD', 'BND', 'OMR'}   # not in DB or zero-stock both days

AM_CASHIER = "jas"
PM_CASHIER = "ana"

# April 5 closing weighted-avg rates (STOCKSLEFT) — used as daily_avg_cost
APR5_AVG = {
    'USD': 59.6,   'JPY': 0.3706, 'KRW': 0.04,
    'AED': 14.98,  'AUD': 40.85,  'CAD': 41.1,   'CHF': 74.65,
    'CNY': 7.88,   'EUR': 68.51,  'GBP': 78.12,  'HKD': 7.5,
    'MYR': 14.01,  'TWD': 1.8,    'NZD': 33.17,  'QAR': 14.03,
    'SAR': 15.28,  'SGD': 45.97,  'THB': 1.84,
    'DKK': 5.6,    'IDR': 0.0033, 'INR': 0.45,
    'NOK': 2.37,   'VND': 0.0026,
}

# Opening positions for April 5 = April 4 STOCKSLEFT (skip zero-qty currencies)
APR4_STOCKS = {
    'USD': (17651, 59.55),    'JPY': (1261000, 0.3704), 'KRW': (3685000, 0.04),
    'AED': (9060, 14.97),     'CAD': (1040, 41.1),      'CHF': (100, 74.58),
    'CNY': (12977, 7.87),     'GBP': (2235, 77.98),     'HKD': (18490, 7.49),
    'MYR': (6360, 14.01),     'TWD': (82700, 1.8),      'NZD': (4310, 33.17),
    'QAR': (847, 14.03),      'SAR': (1228, 15.32),     'SGD': (976, 45.96),
    'THB': (4540, 1.84),      'DKK': (550, 5.6),        'IDR': (410000, 0.0032),
    'INR': (72390, 0.45),     'NOK': (4600, 2.37),      'VND': (164100000, 0.0026),
}

# Column pairs (qty_col, rate_col) per currency for 2ND and OTHERS sheets
PAIRS_2ND = [
    ('AED',1,2), ('AUD',3,4), ('CAD',5,6), ('CHF',7,8),
    ('CNY',9,10), ('EUR',11,12), ('GBP',13,14), ('HKD',15,16),
    ('MYR',17,18), ('NTD',19,20), ('NZD',21,22), ('QR',23,24),
    ('SAR',25,26), ('SGD',27,28), ('THB',29,30),
]
PAIRS_OTHERS = [
    ('BHD',1,2), ('BND',3,4), ('DKK',5,6), ('IDR',7,8),
    ('INR',9,10), ('MOP',11,12), ('NOK',13,14), ('OMR',15,16),
    ('TYR',17,18), ('VND',19,20), ('KD',21,22),
]


def norm(code):
    return CODE_MAP.get(str(code).strip(), str(code).strip())


def is_num(v):
    return isinstance(v, (int, float)) and v > 0


def parse_main(ws):
    """BUY/SELL x MAIN: header at row 0, data from row 1. USD(2,3) JPY(5,6) KRW(8,9).
    Stops after 2 consecutive empty rows to skip Ken's carry-in summary block."""
    pairs = [('USD', 2, 3), ('JPY', 5, 6), ('KRW', 8, 9)]
    txns = []
    empty = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        found = False
        for code, qc, rc in pairs:
            q = row[qc] if len(row) > qc else None
            r = row[rc] if len(row) > rc else None
            if is_num(q) and is_num(r):
                txns.append((code, q, r))
                found = True
        empty = 0 if found else empty + 1
        if empty >= 2:
            break
    return txns


def parse_pairs(ws, pairs, data_start):
    """2ND/OTHERS sheets: currency pairs in cols 1–30 (2ND) or 1–22 (OTHERS).
    Stops after 2 consecutive empty rows to skip Ken's carry-in summary block."""
    txns = []
    empty = 0
    for row in list(ws.iter_rows(values_only=True))[data_start:]:
        found = False
        for code, qc, rc in pairs:
            if code in SKIP_CODES:
                continue
            q = row[qc] if len(row) > qc else None
            r = row[rc] if len(row) > rc else None
            if is_num(q) and is_num(r):
                txns.append((norm(code), q, r))
                found = True
        empty = 0 if found else empty + 1
        if empty >= 2:
            break
    return txns


def make_time(i, n, h0, m0, h1, m1):
    """Distribute index i across n slots between (h0:m0) and (h1:m1)."""
    s0 = h0 * 60 + m0
    s1 = h1 * 60 + m1
    m = int(s0 + (s1 - s0) * i / (n - 1)) if n > 1 else s0
    return f"{m // 60:02d}:{m % 60:02d}"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    buy = (
        [(TxnType.BUY, c, q, r) for c, q, r in parse_main(wb["BUY x MAIN"])]
      + [(TxnType.BUY, c, q, r) for c, q, r in parse_pairs(wb["BUY x 2ND"],    PAIRS_2ND,    2)]
      + [(TxnType.BUY, c, q, r) for c, q, r in parse_pairs(wb["BUY  x OTHERS"], PAIRS_OTHERS, 2)]
    )
    sell = (
        [(TxnType.SELL, c, q, r) for c, q, r in parse_main(wb["SELL x MAIN"])]
      + [(TxnType.SELL, c, q, r) for c, q, r in parse_pairs(wb["SELL  x 2ND"],   PAIRS_2ND,    1)]
      + [(TxnType.SELL, c, q, r) for c, q, r in parse_pairs(wb["SELL x OTHERS"], PAIRS_OTHERS, 1)]
    )
    all_txns = buy + sell
    total    = len(all_txns)
    mid      = total // 2

    print(f"Parsed: {len(buy)} buys, {len(sell)} sells → {total} total transactions")

    # Derive DailyRate: modal buy rate per currency; sell rate from actual sells
    buy_rates_raw = {}
    for _, code, _, rate in buy:
        buy_rates_raw.setdefault(code, []).append(rate)
    dr_buy  = {c: Counter(v).most_common(1)[0][0] for c, v in buy_rates_raw.items()}
    dr_sell = {c: rate for _, c, _, rate in sell}

    db = SessionLocal()
    try:
        # ── DailyRate ──────────────────────────────────────────────────────────
        ins = sk = 0
        for code in sorted(set(dr_buy) | set(dr_sell)):
            if db.query(DailyRate).filter_by(date=TARGET_DATE, currency_code=code).first():
                sk += 1
                continue
            br = dr_buy.get(code, 0)
            sr = dr_sell.get(code, round(br * 1.015, 6))
            db.add(DailyRate(date=TARGET_DATE, currency_code=code,
                             buy_rate=br, sell_rate=sr, set_by="skmc"))
            ins += 1
        db.commit()
        print(f"  Rates      — inserted: {ins}, skipped: {sk}")

        # ── DailyPosition ─────────────────────────────────────────────────────
        ins = sk = 0
        for code, (qty, rate) in APR4_STOCKS.items():
            if db.query(DailyPosition).filter_by(date=TARGET_DATE, currency_code=code).first():
                sk += 1
                continue
            db.add(DailyPosition(date=TARGET_DATE, currency_code=code,
                                 carry_in_qty=qty, carry_in_rate=rate))
            ins += 1
        db.commit()
        print(f"  Positions  — inserted: {ins}, skipped: {sk}")

        # ── Transactions ──────────────────────────────────────────────────────
        ins = sk = 0
        for i, (typ, code, qty, rate) in enumerate(all_txns):
            txn_id = f"OR-{i+1:04d}{DATE_SUFFIX}"
            if db.query(Transaction).filter_by(id=txn_id).first():
                sk += 1
                continue

            if i < mid:
                t_str   = make_time(i,     mid,       8, 0, 13,  0)
                cashier = AM_CASHIER
            else:
                t_str   = make_time(i-mid, total-mid, 13, 0, 17, 30)
                cashier = PM_CASHIER

            avg  = APR5_AVG.get(code, rate)
            than = round((rate - avg) * qty, 4) if typ == TxnType.SELL else 0.0

            db.add(Transaction(
                id=txn_id, date=TARGET_DATE, time=t_str,
                type=typ, source=TxnSource.COUNTER,
                currency_code=code, foreign_amt=qty, rate=rate,
                php_amt=round(qty * rate, 2),
                daily_avg_cost=avg, than=than,
                cashier=cashier, customer=None,
            ))
            ins += 1

        db.commit()
        print(f"  Transactions — inserted: {ins}, skipped: {sk}")
        print("Done.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
