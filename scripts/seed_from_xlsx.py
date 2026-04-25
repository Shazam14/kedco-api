"""
General-purpose daily seeder from Ken's xlsx format.

Usage:
  python scripts/seed_from_xlsx.py <YYYY-MM-DD> <xlsx_path> [am_cashier] [pm_cashier] [prev_xlsx_path]

  YYYY-MM-DD    : date to seed
  xlsx_path     : path to this day's xlsx (transactions + STOCKSLEFT for next day carry-in)
  am_cashier    : username for AM shift (default: jas)
  pm_cashier    : username for PM shift (default: ana)
  prev_xlsx_path: if provided, reads carry-in from that file's STOCKSLEFT
                  if omitted, carry-in is read from DB (already seeded by previous run)

  Add --force to wipe existing records for the date before seeding.

Examples:
  # April 1 — carry-in from March 31 xlsx, nuke existing
  python scripts/seed_from_xlsx.py 2026-04-01 excel_files/4.1.26.xlsx jas ana excel_files/3.31.26.xlsx --force

  # April 6 onward — carry-in already in DB from previous run
  python scripts/seed_from_xlsx.py 2026-04-06 excel_files/4.6.26.xlsx jas ana
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from datetime import date, timedelta
from collections import Counter
import openpyxl

from app.core.database import SessionLocal
from app.models.currency import DailyRate, DailyPosition
from app.models.transaction import Transaction, TxnType, TxnSource

CODE_MAP   = {'NTD': 'TWD', 'QR': 'QAR', 'KD': 'KWD'}
SKIP_CODES = {'BHD', 'BND', 'OMR'}  # TYR activated 2026-04-25

PAIRS_2ND = [
    ('AED',1,2), ('AUD',3,4), ('CAD',5,6), ('CHF',7,8),
    ('CNY',9,10), ('EUR',11,12), ('GBP',13,14), ('HKD',15,16),
    ('MYR',17,18), ('NTD',19,20), ('NZD',21,22), ('QR',23,24),
    ('SAR',25,26), ('SGD',27,28), ('THB',29,30),
]
PAIRS_OTHERS = [
    ('BHD',1,2), ('BND',3,4), ('DKK',5,6), ('IDR',7,8),
    ('INR',9,10), ('MOP',11,12), ('NOK',13,14), ('OMR',15,16),
    ('TYR',17,18), ('VND',19,20), ('KD',21,22),
]


def norm(code):
    return CODE_MAP.get(str(code).strip(), str(code).strip())


def is_num(v):
    return isinstance(v, (int, float)) and v > 0


def parse_main(ws, stop_at_gap=True):
    pairs = [('USD', 2, 3), ('JPY', 5, 6), ('KRW', 8, 9)]
    txns, empty = [], 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        found = False
        for code, qc, rc in pairs:
            q = row[qc] if len(row) > qc else None
            r = row[rc] if len(row) > rc else None
            if is_num(q) and is_num(r):
                txns.append((code, q, r))
                found = True
        if stop_at_gap:
            empty = 0 if found else empty + 1
            if empty >= 2:
                break
    return txns


def parse_pairs(ws, pairs, data_start, stop_at_gap=True, max_rows=None):
    txns, empty = [], 0
    for i, row in enumerate(list(ws.iter_rows(values_only=True))[data_start:]):
        if max_rows is not None and i >= max_rows:
            break
        found = False
        for code, qc, rc in pairs:
            if code in SKIP_CODES:
                continue
            q = row[qc] if len(row) > qc else None
            r = row[rc] if len(row) > rc else None
            if is_num(q) and is_num(r):
                txns.append((norm(code), q, r))
                found = True
            elif is_num(q) and not is_num(r):
                # qty present but no rate — excess/windfall (received for free)
                txns.append((norm(code), q, None))
                found = True
        if stop_at_gap:
            empty = 0 if found else empty + 1
            if empty >= 2:
                break
    return txns


def read_stocksleft(wb):
    stocks = {}
    for row in wb['STOCKSLEFT'].iter_rows(values_only=True):
        code, qty, rate = row[0], row[1], row[2]
        if not isinstance(code, str) or len(code.strip()) < 2 or len(code.strip()) > 5:
            continue
        db_code = CODE_MAP.get(code.strip(), code.strip())
        if db_code in SKIP_CODES:
            continue
        if isinstance(qty, (int, float)) and qty > 0 and isinstance(rate, (int, float)) and rate > 0:
            stocks[db_code] = (qty, rate)
    return stocks


def make_time(i, n, h0, m0, h1, m1):
    s0 = h0 * 60 + m0
    s1 = h1 * 60 + m1
    m = int(s0 + (s1 - s0) * i / (n - 1)) if n > 1 else s0
    return f"{m // 60:02d}:{m % 60:02d}"


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    force = '--force' in sys.argv

    if len(args) < 2:
        print("Usage: seed_from_xlsx.py <YYYY-MM-DD> <xlsx> [am_cashier] [pm_cashier] [prev_xlsx]")
        sys.exit(1)

    target_date = date.fromisoformat(args[0])
    xlsx_path   = args[1]
    am_cashier  = args[2] if len(args) > 2 else 'jas'
    pm_cashier  = args[3] if len(args) > 3 else 'ana'
    prev_xlsx   = args[4] if len(args) > 4 else None

    date_suffix = target_date.strftime('%m%d')
    next_date   = target_date + timedelta(days=1)

    print(f"Seeding {target_date} from {xlsx_path}")
    print(f"  Cashiers: AM={am_cashier}, PM={pm_cashier}")
    print(f"  Carry-in: {'from ' + prev_xlsx if prev_xlsx else 'from DB'}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def ws(name):
        """Return worksheet by name, tolerating case/spacing/typo variants (e.g. 'BU X MAIN')."""
        if name in wb.sheetnames:
            return wb[name]
        # Normalise both sides: lowercase, strip, collapse whitespace
        def norm(s):
            return ' '.join(s.lower().split())
        n = norm(name)
        for s in wb.sheetnames:
            if norm(s) == n:
                return wb[s]
        # Handle 'BU' typo for 'BUY' (and vice-versa)
        n2 = n.replace('buy', 'bu')
        for s in wb.sheetnames:
            if norm(s).replace('buy', 'bu') == n2:
                return wb[s]
        raise KeyError(f"Worksheet '{name}' not found in {wb.sheetnames}")

    # When prev_xlsx is provided, carry-in rows are stripped explicitly so we can
    # parse the full sheet (no early exit). Without prev_xlsx, the 2-empty-row stop
    # naturally excludes the carry-in block at the end.
    stop = prev_xlsx is None

    def classify_buy(c, q, r):
        """BUY when rate is present; EXCESS when qty exists but rate is missing."""
        return (TxnType.EXCESS, c, q, 0) if r is None else (TxnType.BUY, c, q, r)

    buy = (
        [classify_buy(c, q, r) for c, q, r in parse_main(ws('BUY x MAIN'), stop)]
      + [classify_buy(c, q, r) for c, q, r in parse_pairs(ws('BUY x 2ND'),    PAIRS_2ND,    2, stop)]
      + [classify_buy(c, q, r) for c, q, r in parse_pairs(ws('BUY  x OTHERS'), PAIRS_OTHERS, 2, stop)]
    )
    sell = (
        [(TxnType.SELL, c, q, r) for c, q, r in parse_main(ws('SELL x MAIN'), True)]
      + [(TxnType.SELL, c, q, r) for c, q, r in parse_pairs(ws('SELL  x 2ND'),   PAIRS_2ND,    1, True)]
      + [(TxnType.SELL, c, q, r) for c, q, r in parse_pairs(ws('SELL x OTHERS'), PAIRS_OTHERS, 1, True)]
    )

    # Carry-in: from prev xlsx STOCKSLEFT or from DB
    if prev_xlsx:
        prev_wb  = openpyxl.load_workbook(prev_xlsx, data_only=True)
        carry_in = read_stocksleft(prev_wb)
        prev_wb.close()
    else:
        carry_in = None  # will read from DB per currency below

    # Each BUY sheet has the previous day's carry-in as a summary row (same qty/rate as
    # DailyPosition). Strip any BUY that exactly matches a carry_in entry so it isn't
    # double-counted. When prev_xlsx is absent, fall back to existing DB positions.
    if carry_in:
        carry_in_set = {(code, qty, rate) for code, (qty, rate) in carry_in.items()}
    else:
        _db = SessionLocal()
        try:
            _pos = _db.query(DailyPosition).filter_by(date=target_date).all()
            carry_in_set = {(p.currency_code, p.carry_in_qty, p.carry_in_rate) for p in _pos}
            # Preserve carry_in so --force can recreate positions after wiping them
            carry_in = {p.currency_code: (p.carry_in_qty, p.carry_in_rate) for p in _pos}
        finally:
            _db.close()

    if carry_in_set:
        buy_clean = [(typ, c, q, r) for typ, c, q, r in buy if (c, q, r) not in carry_in_set]
        removed = len(buy) - len(buy_clean)
        if removed:
            print(f"  Stripped {removed} carry-in BUY row(s) from parsed transactions")
        buy = buy_clean

    all_txns = buy + sell
    total    = len(all_txns)
    mid      = total // 2

    print(f"  Parsed: {len(buy)} buys, {len(sell)} sells → {total} total")

    # Derive DailyRate: modal buy rate per currency; actual sell rate from sells
    buy_rates_raw = {}
    for _, code, _, rate in buy:
        buy_rates_raw.setdefault(code, []).append(rate)
    dr_buy  = {c: Counter(v).most_common(1)[0][0] for c, v in buy_rates_raw.items()}
    dr_sell = {c: rate for _, c, _, rate in sell}

    # Closing avg per currency from this day's STOCKSLEFT — used for THAN on sells
    today_stocks = read_stocksleft(wb)
    daily_avg    = {code: rate for code, (qty, rate) in today_stocks.items()}

    wb.close()

    db = SessionLocal()
    try:
        if force:
            deleted = db.query(Transaction).filter_by(date=target_date).delete()
            db.query(DailyRate).filter_by(date=target_date).delete()
            db.query(DailyPosition).filter_by(date=target_date).delete()
            db.commit()
            print(f"  Cleared {deleted} transactions + rates + positions for {target_date}")

        # ── DailyRate ──────────────────────────────────────────────────────────
        ins = sk = 0
        for code in sorted(set(dr_buy) | set(dr_sell)):
            if db.query(DailyRate).filter_by(date=target_date, currency_code=code).first():
                sk += 1; continue
            br = dr_buy.get(code, 0)
            sr = dr_sell.get(code, round(br * 1.015, 6))
            db.add(DailyRate(date=target_date, currency_code=code,
                             buy_rate=br, sell_rate=sr, set_by='skmc'))
            ins += 1
        db.commit()
        print(f"  Rates      — inserted: {ins}, skipped: {sk}")

        # ── DailyPosition (today's carry-in) ──────────────────────────────────
        ins = sk = 0
        if carry_in:
            for code, (qty, rate) in carry_in.items():
                if db.query(DailyPosition).filter_by(date=target_date, currency_code=code).first():
                    sk += 1; continue
                db.add(DailyPosition(date=target_date, currency_code=code,
                                     carry_in_qty=qty, carry_in_rate=rate))
                ins += 1
            db.commit()
            print(f"  Positions  — inserted: {ins}, skipped: {sk}")
        else:
            print(f"  Positions  — using existing DB carry-in")

        # ── DailyPosition next day (carry-in from today's STOCKSLEFT) ─────────
        ins = sk = 0
        for code, (qty, rate) in today_stocks.items():
            if db.query(DailyPosition).filter_by(date=next_date, currency_code=code).first():
                sk += 1; continue
            db.add(DailyPosition(date=next_date, currency_code=code,
                                 carry_in_qty=qty, carry_in_rate=rate))
            ins += 1
        db.commit()
        print(f"  Positions+1 ({next_date}) — inserted: {ins}, skipped: {sk}")

        # ── Transactions ──────────────────────────────────────────────────────
        ins = sk = 0
        for i, (typ, code, qty, rate) in enumerate(all_txns):
            txn_id = f"OR-{i+1:04d}{date_suffix}"
            if db.query(Transaction).filter_by(id=txn_id).first():
                sk += 1; continue

            if i < mid:
                t_str   = make_time(i,     mid,       8, 0, 13,  0)
                cashier = am_cashier
            else:
                t_str   = make_time(i-mid, total-mid, 13, 0, 17, 30)
                cashier = pm_cashier

            if typ == TxnType.EXCESS:
                db.add(Transaction(
                    id=txn_id, date=target_date, time=t_str,
                    type=typ, source=TxnSource.COUNTER,
                    currency_code=code, foreign_amt=qty, rate=0.0,
                    php_amt=0.0, daily_avg_cost=0.0, than=0.0,
                    cashier=cashier, customer=None,
                    note='Excess received — no rate (from Excel)',
                ))
            else:
                avg  = daily_avg.get(code, rate)
                than = round((rate - avg) * qty, 4) if typ == TxnType.SELL else 0.0
                db.add(Transaction(
                    id=txn_id, date=target_date, time=t_str,
                    type=typ, source=TxnSource.COUNTER,
                    currency_code=code, foreign_amt=qty, rate=rate,
                    php_amt=round(qty * rate, 2),
                    daily_avg_cost=avg, than=than,
                    cashier=cashier, customer=None,
                ))
            ins += 1

        db.commit()
        print(f"  Transactions — inserted: {ins}, skipped: {sk}")
        print("Done.")
    finally:
        db.close()


if __name__ == '__main__':
    main()
